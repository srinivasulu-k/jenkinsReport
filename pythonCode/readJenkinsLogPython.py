'''
Created on March 23, 2021

@author: Srinivasulu
'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
import openpyxl 
import sys
import string
import datetime
# from datetime import date
import os
import json

testPlanName = sys.argv[1]
buildURL = sys.argv[2]
buildNumber = sys.argv[3]

# testPlanName="DAM Smoke Tests UAT"
# buildURL="http://na1qalabd1:8080/job/DAM%20Smoke%20Tests%20UAT/lastBuild/"
# buildNumber=0

# today = date.today()
# dayOfWeek = today.weekday()
# print("Test Plan Name: "+testPlanName)
# print("Build URL: "+buildURL)

print("Execution Started")
driver=webdriver.Chrome(executable_path=str(os.getcwd())+'\\webDriver\\chromedriver.exe')
driver.implicitly_wait(10)
driver.maximize_window()
driver.get( buildURL + "timestamps/?elapsed=HH:mm:ss.S" )

txtTimeContent=str(driver.find_element_by_tag_name("pre").text)
txtTimeElapsed=txtTimeContent.splitlines()[-1]

# print(txtTimeContent)
# print("Total Time: "+txtTimeElapsed)

driver.get( buildURL + "logText/progressiveText?start=0" )

txtContent=str(driver.find_element_by_tag_name("pre").text)
txtContentLastLine=txtContent.splitlines()[-1]
txtBuildStatus=txtContentLastLine.split()[-1]
# print("Build Status: "+txtBuildStatus)

buildDate1=(txtContent.split("Finished at: "))
buildDate=(buildDate1[1].split("T"))
buildDate=buildDate[0]

featureFileList=(txtContent.split("Feature file file:"))
featureFileSet=set()
for i in featureFileList:
    featureName=(i.split("\n")[0].split(" *****")[0]).split(".feature:")[0]
    featureFileSet.add(featureName)

# print("Set Data: ")
# print(featureFileSet)
featureFileCount=int((len(featureFileSet)) - 1)
# print("Total Features: " + str(featureFileCount))

resultsLine=(txtContent.split("run:")[-1])
# print("Line- "+resultsLine)

resultsLine=resultsLine.split(", ")
totalTests=(resultsLine[0]).split(" ")[-1]
totalFailures=(resultsLine[1]).split(" ")[-1]
totalError=(resultsLine[2]).split(" ")[-1]
totalSkipped=((resultsLine[3]).split("\n")[0]).split(" ")[-1]
totalPassed=(int(totalTests) - (int(totalFailures) + int(totalError) + int(totalSkipped)))

driver.close()

# print("Total Tests--"+totalTests+"--")
# print("Total Failures--"+totalFailures+"--")
# print("Total Errors--"+totalError+"--")
# print("Total Skipped--"+totalSkipped+"--")

# filename1 = 'C:\\wamp64\\www\\jenkinsReport\\dataFiles\\testExcel.xlsx'
filename1 = str(os.getcwd())+'\\dataFiles\\testExcel.xlsx'
wb = openpyxl.load_workbook(filename1)
ws = wb.active
rowNum=ws.max_row+1

ws.cell(column=1, row=rowNum, value=testPlanName)
ws.cell(column=2, row=rowNum, value=int(buildNumber))
ws.cell(column=3, row=rowNum, value=buildDate)
ws.cell(column=4, row=rowNum, value=txtTimeElapsed)
ws.cell(column=5, row=rowNum, value=txtBuildStatus)
ws.cell(column=6, row=rowNum, value=int(featureFileCount))
ws.cell(column=7, row=rowNum, value=int(totalTests))
ws.cell(column=8, row=rowNum, value=int(totalPassed))
ws.cell(column=9, row=rowNum, value=int(totalFailures))
ws.cell(column=10, row=rowNum, value=int(totalError))
ws.cell(column=11, row=rowNum, value=int(totalSkipped))
wb.save(filename1)

def write_json(data,filename=str((os.getcwd())+'\\dataFiles\\testJSON.json')):
    with open (filename, "w") as f:
        json.dump(data, f, indent=4)

with open (str(os.getcwd())+'\\dataFiles\\testJSON.json') as json_file:
    data = json.load(json_file)
    temp = data["damJenkinsJobs"]
    # temp = data[testPlanName.replace(" " , "")]
    y = {"testPlanName" : testPlanName,
        "buildNumber" : int(buildNumber),
        "buildDate" : buildDate,
        "txtTimeElapsed" : txtTimeElapsed,
        "txtBuildStatus" : txtBuildStatus,
        "featureFileCount" : int(featureFileCount),
        "totalTests" : int(totalTests),
        "totalPassed" : int(totalPassed),
        "totalFailures" : int(totalFailures),
        "totalError" : int(totalError),
        "totalSkipped" : int(totalSkipped)
        }
    temp.append(y)

write_json(data)

print("Execution Completed")
